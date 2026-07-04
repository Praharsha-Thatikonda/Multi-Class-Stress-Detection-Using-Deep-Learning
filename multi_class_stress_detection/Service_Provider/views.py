from django.db.models import Count, Avg
from django.shortcuts import render, redirect
from django.db.models import Q
import datetime
import io
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
import numpy as np

import pandas as pd  # data processing, CSV file I/O (e.g. pd.read_csv)
from sklearn.feature_extraction.text import CountVectorizer

from sklearn.tree import DecisionTreeClassifier

# model selection
from sklearn.metrics import (
    confusion_matrix,
    accuracy_score,
    ConfusionMatrixDisplay,
    classification_report,
)

# Create your views here.
from Remote_User.models import (
    ClientRegister_Model,
    predict_stress_detection,
    detection_ratio,
    detection_accuracy,
)


def serviceproviderlogin(request):
    if request.method == "POST":
        admin = request.POST.get('username')
        password = request.POST.get('password')
        if admin == "Admin" and password == "Admin":
            detection_accuracy.objects.all().delete()
            return redirect('View_Remote_Users')

    return render(request, 'SProvider/serviceproviderlogin.html')


def View_Predict_Stress_Detection_Type_Ratio(request):
    detection_ratio.objects.all().delete()

    kword = 'Stress'
    print(kword)
    obj = predict_stress_detection.objects.filter(Prediction=kword)
    total = predict_stress_detection.objects.count()

    if total > 0:
        count = obj.count()
        ratio = (count / total) * 100
        if ratio != 0:
            detection_ratio.objects.create(names=kword, ratio=ratio)

        kword1 = 'No Stress'
        print(kword1)
        obj1 = predict_stress_detection.objects.filter(Prediction=kword1)
        count1 = obj1.count()
        ratio1 = (count1 / total) * 100
        if ratio1 != 0:
            detection_ratio.objects.create(names=kword1, ratio=ratio1)

    obj = detection_ratio.objects.all()
    return render(request, 'SProvider/View_Predict_Stress_Detection_Type_Ratio.html', {'objs': obj})


def View_Remote_Users(request):
    obj = ClientRegister_Model.objects.all()
    return render(request, 'SProvider/View_Remote_Users.html', {'objects': obj})


def ViewTrendings(request):
    topic = predict_stress_detection.objects.values('topics').annotate(
        dcount=Count('topics')
    ).order_by('-dcount')
    return render(request, 'SProvider/ViewTrendings.html', {'objects': topic})


def charts(request, chart_type):
    chart1 = detection_ratio.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request, "SProvider/charts.html", {'form': chart1, 'chart_type': chart_type})


def charts1(request, chart_type):
    chart1 = detection_accuracy.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request, "SProvider/charts1.html", {'form': chart1, 'chart_type': chart_type})


def View_Predict_Stress_Detection_Details(request):
    obj = predict_stress_detection.objects.all()
    return render(request, 'SProvider/View_Predict_Stress_Detection_Details.html', {'list_objects': obj})


def likeschart(request, like_chart):
    charts = detection_accuracy.objects.values('names').annotate(dcount=Avg('ratio'))
    return render(request, "SProvider/likeschart.html", {'form': charts, 'like_chart': like_chart})


def Download_Trained_DataSets(request):
    """
    Download all predicted stress detection records as an Excel (.xlsx) file.
    Uses openpyxl instead of the deprecated/unsupported xlwt library.
    """
    # Column headers
    headers = [
        'FID', 'MEAN_RR', 'MEDIAN_RR', 'SDRR', 'RMSSD', 'SDSD',
        'SDRR_RMSSD', 'HR', 'VLF', 'VLF_PCT', 'LF', 'LF_PCT', 'LF_NU',
        'HF', 'HF_PCT', 'HF_NU', 'TP', 'LF_HF', 'HF_LF', 'sampen',
        'higuci', 'Prediction',
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predicted Data"

    bold_font = Font(bold=True)

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Write data rows
    obj = predict_stress_detection.objects.all()
    for row_idx, my_row in enumerate(obj, start=2):
        values = [
            my_row.FID, my_row.MEAN_RR, my_row.MEDIAN_RR, my_row.SDRR,
            my_row.RMSSD, my_row.SDSD, my_row.SDRR_RMSSD, my_row.HR,
            my_row.VLF, my_row.VLF_PCT, my_row.LF, my_row.LF_PCT,
            my_row.LF_NU, my_row.HF, my_row.HF_PCT, my_row.HF_NU,
            my_row.TP, my_row.LF_HF, my_row.HF_LF, my_row.sampen,
            my_row.higuci, my_row.Prediction,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to in-memory buffer and return as HTTP response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Predicted_Data.xlsx"'
    return response


def train_model(request):
    detection_accuracy.objects.all().delete()
    df = pd.read_csv('Datasets.csv', encoding='latin-1')

    def apply_results(label):
        if label == 'no stress':
            return 0  # No Stress
        elif label == 'stress':
            return 1  # Stress
        return None

    df['results'] = df['condition'].apply(apply_results)

    x = df["FID"]
    y = df["results"]

    cv = CountVectorizer(lowercase=False, strip_accents='unicode', ngram_range=(1, 1))
    x = cv.fit_transform(x)

    labeled = 'Results_data.csv'
    df.to_csv(labeled, index=False)

    print("Data")
    print(x)
    print("Results")
    print(y)

    models = []
    from sklearn.model_selection import train_test_split
    X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.30, random_state=123)
    print(X_train.shape, X_test.shape, y_train.shape)

    print("Convolution Neural Network (CNN)")
    from sklearn.neural_network import MLPClassifier
    mlpc = MLPClassifier(max_iter=500).fit(X_train, y_train)
    y_pred = mlpc.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(y_test, y_pred) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(y_test, y_pred))
    print("CONFUSION MATRIX")
    print(confusion_matrix(y_test, y_pred))
    models.append(('MLPClassifier', mlpc))
    detection_accuracy.objects.create(
        names="Convolution Neural Network (CNN)",
        ratio=accuracy_score(y_test, y_pred) * 100,
    )

    # SVM Model
    print("SVM")
    from sklearn import svm
    lin_clf = svm.LinearSVC()
    lin_clf.fit(X_train, y_train)
    predict_svm = lin_clf.predict(X_test)
    svm_acc = accuracy_score(y_test, predict_svm) * 100
    print("ACCURACY")
    print(svm_acc)
    print("CLASSIFICATION REPORT")
    print(classification_report(y_test, predict_svm))
    print("CONFUSION MATRIX")
    print(confusion_matrix(y_test, predict_svm))
    detection_accuracy.objects.create(names="SVM", ratio=svm_acc)

    print("Logistic Regression")
    from sklearn.linear_model import LogisticRegression
    reg = LogisticRegression(random_state=0, solver='lbfgs', max_iter=1000).fit(X_train, y_train)
    y_pred = reg.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(y_test, y_pred) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(y_test, y_pred))
    print("CONFUSION MATRIX")
    print(confusion_matrix(y_test, y_pred))
    detection_accuracy.objects.create(
        names="Logistic Regression",
        ratio=accuracy_score(y_test, y_pred) * 100,
    )

    print("Decision Tree Classifier")
    dtc = DecisionTreeClassifier()
    dtc.fit(X_train, y_train)
    dtcpredict = dtc.predict(X_test)
    print("ACCURACY")
    print(accuracy_score(y_test, dtcpredict) * 100)
    print("CLASSIFICATION REPORT")
    print(classification_report(y_test, dtcpredict))
    print("CONFUSION MATRIX")
    print(confusion_matrix(y_test, dtcpredict))
    detection_accuracy.objects.create(
        names="Decision Tree Classifier",
        ratio=accuracy_score(y_test, dtcpredict) * 100,
    )

    df.to_csv(labeled, index=False)

    obj = detection_accuracy.objects.all()
    return render(request, 'SProvider/train_model.html', {'objs': obj})